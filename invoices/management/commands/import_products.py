from django.core.management.base import BaseCommand
from invoices.models import Product, ProductCategory, Unit, ChartOfAccount
from openpyxl import load_workbook
from decimal import Decimal
import os

class Command(BaseCommand):
    help = "استيراد بيانات المنتجات من ملف Excel إلى قاعدة البيانات"

    def handle(self, *args, **options):
        # تحديد مسار ملف Excel
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        excel_path = os.path.join(base_dir, 'products.xlsx')
        
        # فتح ملف Excel وقراءة الورقة النشطة
        wb = load_workbook(filename=excel_path)
        ws = wb.active

        # التحقق من وجود بيانات (تخطي الصف الأول الذي يحتوي على الرؤوس)
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            name_ar = row[0]
            serial_number = row[1]
            category_name = row[2]
            unit_name = row[3]
            price = row[4]
            description = row[5]
            stock = row[6]
            low_stock_threshold = row[7]
            inventory_account_name = row[8]
            
            # البحث عن الكائنات المرتبطة
            try:
                category = ProductCategory.objects.get(name=category_name)
            except ProductCategory.DoesNotExist:
                category = None
            
            try:
                unit = Unit.objects.get(name=unit_name)
            except Unit.DoesNotExist:
                unit = None
            
            try:
                inventory_account = ChartOfAccount.objects.get(name=inventory_account_name)
            except ChartOfAccount.DoesNotExist:
                inventory_account = None
            
            # إنشاء سجل المنتج
            Product.objects.create(
                name_ar=name_ar,
                serial_number=serial_number,
                category=category,
                unit=unit,
                price=Decimal(price) if price is not None else Decimal('0.00'),
                description=description,
                stock=int(stock) if stock is not None else 999999999,
                low_stock_threshold=int(low_stock_threshold) if low_stock_threshold is not None else 10,
                inventory_account=inventory_account,
            )

        self.stdout.write(self.style.SUCCESS(f"تم استيراد {row_count} منتج بنجاح."))
